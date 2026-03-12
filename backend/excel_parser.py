"""
Excel parser for Annexure 2 (benchmarking workings) and Annexure 3 (tested party margins).
Extracts structured data from Excel sheets for document generation.
"""

import openpyxl
from typing import Optional


class SearchStrategyRow:
    def __init__(self, criterion: str, description: str, count: str):
        self.criterion = criterion
        self.description = description
        self.count = count


class SearchStrategyData:
    def __init__(self, region: str, rows: list[SearchStrategyRow], result_count: int):
        self.region = region
        self.rows = rows
        self.result_count = result_count


class RejectionReason:
    def __init__(self, number: int, reason: str, count: int):
        self.number = number
        self.reason = reason
        self.count = count


class ComparableCompany:
    def __init__(self, sno: int, name: str, margins: dict[str, Optional[str]], weighted_avg: str):
        self.sno = sno
        self.name = name
        self.margins = margins  # e.g. {"FY 2022": "1.35%", "FY 2023": "2.96%", ...}
        self.weighted_avg = weighted_avg


class QuartileData:
    def __init__(self, lower_quartile: str, median: str, upper_quartile: str,
                 per_year_lower: dict = None, per_year_median: dict = None,
                 per_year_upper: dict = None):
        self.lower_quartile = lower_quartile
        self.median = median
        self.upper_quartile = upper_quartile
        self.per_year_lower = per_year_lower or {}   # {"FY 2022": "1.80%", ...}
        self.per_year_median = per_year_median or {}
        self.per_year_upper = per_year_upper or {}


class MarginAnalysisCompany:
    def __init__(self, sno: int, name: str, country: str, bvd_id: str,
                 ec_oc: dict[str, Optional[str]], weighted_ec_oc: str,
                 markup: dict[str, Optional[str]], weighted_markup: str,
                 margin_revenue: dict[str, Optional[str]], weighted_margin: str):
        self.sno = sno
        self.name = name
        self.country = country
        self.bvd_id = bvd_id
        self.ec_oc = ec_oc
        self.weighted_ec_oc = weighted_ec_oc
        self.markup = markup
        self.weighted_markup = weighted_markup
        self.margin_revenue = margin_revenue
        self.weighted_margin = weighted_margin


class AcceptRejectEntry:
    def __init__(self, company_name: str, result: str):
        self.company_name = company_name
        self.result = result


class BenchmarkingData:
    """All data extracted from Annexure 2 Excel"""
    def __init__(self):
        self.search_strategies: list[SearchStrategyData] = []
        self.rejection_reasons: list[RejectionReason] = []
        self.total_rejections: int = 0
        self.comparable_companies: list[ComparableCompany] = []
        self.quartiles: Optional[QuartileData] = None
        self.fy_years: list[str] = []  # e.g. ["FY 2022", "FY 2023", "FY 2024"]
        self.total_accepted: int = 0
        self.accept_reject_entries: list[AcceptRejectEntry] = []
        self.margin_analysis: list[MarginAnalysisCompany] = []


def _safe_str(val) -> str:
    """Convert cell value to string safely."""
    if val is None:
        return ""
    return str(val).strip()


def _safe_pct(val) -> str:
    """Convert a decimal or percentage value to formatted percentage string."""
    if val is None:
        return "N/A"
    if isinstance(val, str):
        val = val.strip()
        if val.upper() == "N/A" or val == "" or val == "-":
            return "N/A"
        if "%" in val:
            return val
        try:
            val = float(val)
        except ValueError:
            return val
    if isinstance(val, (int, float)):
        if abs(val) < 1:
            return f"{val * 100:.2f}%"
        else:
            return f"{val:.2f}%"
    return str(val)


def _parse_search_strategy_sheet(ws, region: str) -> SearchStrategyData:
    """Parse a search strategy sheet. Data starts at column B (index 1)."""
    rows = []
    result_count = 0

    data_started = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [_safe_str(c) for c in row]

        # Data is at columns B(1), C(2), D(3)
        col1 = cells[1] if len(cells) > 1 else ""
        col2 = cells[2] if len(cells) > 2 else ""
        col3 = cells[3] if len(cells) > 3 else ""

        if not col1:
            continue

        # Skip header rows until we find actual numbered criteria
        if not data_started:
            if col1.strip() == "Search Strategy":
                data_started = True
                continue
            if "Tour" in col1 or "Annexure" in col1 or "Search Strategy" in col1:
                continue
            continue

        # Parse data rows - criteria start with a number or "Boolean"
        criterion = col1.strip()
        if not criterion:
            continue

        # Check if this is a numbered criterion or Boolean search
        is_criterion = (
            any(criterion.startswith(f"{i}.") for i in range(1, 15)) or
            criterion.startswith("Boolean") or
            criterion.startswith("\t")  # Some criteria have tabs
        )

        if is_criterion:
            criterion = criterion.strip("\t").strip()
            description = col2.strip() if col2 else ""
            count = col3.strip() if col3 else ""

            rows.append(SearchStrategyRow(criterion, description, count))
            if "Boolean" in criterion:
                try:
                    result_count = int(count)
                except (ValueError, TypeError):
                    result_count = 0

    return SearchStrategyData(region=region, rows=rows, result_count=result_count)


def _parse_accept_reject_matrix(ws) -> tuple[list[RejectionReason], int, list[AcceptRejectEntry], int]:
    """Parse the accept/reject matrix sheet. Data starts at column B (index 1)."""
    rejection_reasons = []
    accept_reject_entries = []
    total_rejections = 0
    total_accepted = 0

    # First section: summary counts
    in_summary = True
    reason_number = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [_safe_str(c) for c in row]

        if not any(cells):
            continue

        # Data is in column B (index 1) and C (index 2)
        col1 = cells[1] if len(cells) > 1 else ""
        col2 = cells[2] if len(cells) > 2 else ""

        # Skip header rows — match common activity/service names and labels
        if col1 in ("Tour Operation services", "Accept Reject Matrix", "Annexure 9", ""):
            continue
        # Skip any row where col1 looks like an activity header (no count or count is 0)
        if col2 == "" or col2 == "0":
            if col1 not in ("Accept", "Total") and "Reject" not in col1:
                continue

        if col1 == "Accept/ Reject Reason" and col2 == "Count":
            continue

        if in_summary:
            if col1 == "Company Name":
                in_summary = False
                continue

            if col1 == "Total":
                continue

            reason = col1
            count_str = col2

            if reason == "Accept":
                try:
                    total_accepted = int(count_str)
                except (ValueError, TypeError):
                    total_accepted = 0
            elif reason:
                reason_number += 1
                try:
                    count = int(count_str)
                except (ValueError, TypeError):
                    count = 0
                rejection_reasons.append(RejectionReason(reason_number, reason, count))
                total_rejections += count
        else:
            # Individual company entries
            if col1 and col1 not in ("Company Name",):
                accept_reject_entries.append(AcceptRejectEntry(col1, col2))

    return rejection_reasons, total_rejections, accept_reject_entries, total_accepted


def _parse_margin_analysis(ws) -> tuple[list[MarginAnalysisCompany], list[str], QuartileData]:
    """Parse the Margin Analysis sheet. Data starts at column B (index 1)."""
    companies = []
    fy_years = []
    quartile_data = None

    # Column layout (1-indexed from B column):
    # B=S.No, C=Name, D=Country, E=BvD ID,
    # F-H=EC(FY22,23,24), I-K=OR(FY22,23,24), L-N=OP(FY22,23,24), O-Q=OC(FY22,23,24),
    # R-T=EC/OC(FY22,23,24), U=w_EC/OC, V-X=Markup(FY22,23,24), Y=w_Markup,
    # Z-AB=Margin(FY22,23,24), AC=w_Margin
    # In 0-indexed terms: B=1, so margin cols are at 25,26,27,28

    # Find the FY row to extract year labels
    fy_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        cells = [_safe_str(c) for c in row]
        if any("FY" in c for c in cells if c):
            fy_row = row_idx + 1
            for c in cells:
                if c and c.startswith("FY ") and c not in fy_years:
                    fy_years.append(c)
            break

    # Deduplicate FY years (they appear multiple times in header)
    seen = []
    for fy in fy_years:
        if fy not in seen:
            seen.append(fy)
    fy_years = seen[:3]  # Only keep unique FY years (not the weighted avg label)

    if not fy_years:
        fy_years = ["FY 2022", "FY 2023", "FY 2024"]

    data_start = (fy_row or 7) + 1  # Start after FY header row

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=True):
        cells = list(row)
        if not cells or len(cells) < 5:
            continue

        # S.No is at column B (index 1)
        cell1 = cells[1] if len(cells) > 1 else None
        cell1_str = _safe_str(cell1)

        if not cell1_str:
            continue

        # Skip non-data rows
        if cell1_str in ("Tour Operation services", "Margin Analysis", "Annexure 11", "S. No.", ""):
            continue

        try:
            sno_val = int(cell1)
        except (ValueError, TypeError):
            continue

        name = _safe_str(cells[2]) if len(cells) > 2 else ""
        country = _safe_str(cells[3]) if len(cells) > 3 else ""
        bvd_id = _safe_str(cells[4]) if len(cells) > 4 else ""

        # Margin on total revenue is the last 4 non-empty columns
        # Based on structure: cols 25,26,27 = FY margins, col 28 = weighted avg
        total_cols = len(cells)
        margin_cols_start = total_cols - 4  # Last 4 columns

        margin_revenue = {}
        for i, fy in enumerate(fy_years[:3]):
            col_idx = margin_cols_start + i
            if col_idx < total_cols:
                margin_revenue[fy] = _safe_pct(cells[col_idx])
            else:
                margin_revenue[fy] = "N/A"

        weighted_margin = _safe_pct(cells[total_cols - 1]) if total_cols > 0 else "N/A"

        companies.append(MarginAnalysisCompany(
            sno=sno_val, name=name, country=country, bvd_id=bvd_id,
            ec_oc={}, weighted_ec_oc="N/A",
            markup={}, weighted_markup="N/A",
            margin_revenue=margin_revenue, weighted_margin=weighted_margin
        ))

    quartile_data = QuartileData("", "", "")
    return companies, fy_years, quartile_data


def parse_annexure2(file_path: str) -> BenchmarkingData:
    """
    Parse the Annexure 2 Excel workbook and extract all benchmarking data.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = BenchmarkingData()

    # Parse search strategy sheets
    strategy_sheets = {
        "UAE Search Strategy": "UAE",
        "Mena & Turkiye Search Strategy": "MENA & Turkey",
        "Europe Search Strategy": "Eastern Europe, MENA, Turkey & UAE",
    }

    for sheet_name, region in strategy_sheets.items():
        if sheet_name in wb.sheetnames:
            ss = _parse_search_strategy_sheet(wb[sheet_name], region)
            data.search_strategies.append(ss)

    # Parse accept/reject matrix
    if "Accept reject matrix" in wb.sheetnames:
        ws = wb["Accept reject matrix"]
        data.rejection_reasons, data.total_rejections, data.accept_reject_entries, data.total_accepted = \
            _parse_accept_reject_matrix(ws)

    # Parse margin analysis
    if "Margin Analysis" in wb.sheetnames:
        ws = wb["Margin Analysis"]
        data.margin_analysis, data.fy_years, _ = _parse_margin_analysis(ws)

    # Parse summary sheet for quartile data
    if "Summary of TP BM" in wb.sheetnames:
        ws = wb["Summary of TP BM"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            cells = [_safe_str(c) for c in row]
            raw_cells = list(row)  # Keep raw numeric values

            if any("Lower quartile" in c for c in cells if c):
                # OP/OR columns: index 6=FY2022, 7=FY2023, 8=FY2024, 9=weighted avg
                # (Summary sheet layout: B=PLI, C-F=EC/OC, G-J=OP/OR)
                op_or_vals = {}
                weighted = ""
                fy_names = data.fy_years if data.fy_years else ["FY 2022", "FY 2023", "FY 2024"]
                for ci in range(len(raw_cells)):
                    val = raw_cells[ci]
                    if val is not None and isinstance(val, (int, float)) and abs(val) < 1:
                        # Map column index to FY year
                        if ci >= 6 and ci <= 8:
                            fy_idx = ci - 6
                            if fy_idx < len(fy_names):
                                op_or_vals[fy_names[fy_idx]] = _safe_pct(val)
                        if ci == 9:
                            weighted = _safe_pct(val)
                if not weighted:
                    # Fall back: use last float < 1
                    all_floats = [v for v in raw_cells if v is not None and isinstance(v, (int, float)) and abs(v) < 1]
                    weighted = _safe_pct(all_floats[-1]) if all_floats else "N/A"
                data.quartiles = QuartileData(
                    lower_quartile=weighted, median="", upper_quartile="",
                    per_year_lower=op_or_vals
                )

            if any("Median" in c for c in cells if c):
                op_or_vals = {}
                weighted = ""
                fy_names = data.fy_years if data.fy_years else ["FY 2022", "FY 2023", "FY 2024"]
                for ci in range(len(raw_cells)):
                    val = raw_cells[ci]
                    if val is not None and isinstance(val, (int, float)) and abs(val) < 1:
                        if ci >= 6 and ci <= 8:
                            fy_idx = ci - 6
                            if fy_idx < len(fy_names):
                                op_or_vals[fy_names[fy_idx]] = _safe_pct(val)
                        if ci == 9:
                            weighted = _safe_pct(val)
                if not weighted:
                    all_floats = [v for v in raw_cells if v is not None and isinstance(v, (int, float)) and abs(v) < 1]
                    weighted = _safe_pct(all_floats[-1]) if all_floats else "N/A"
                if data.quartiles:
                    data.quartiles.median = weighted
                    data.quartiles.per_year_median = op_or_vals

            if any("Upper quartile" in c for c in cells if c):
                op_or_vals = {}
                weighted = ""
                fy_names = data.fy_years if data.fy_years else ["FY 2022", "FY 2023", "FY 2024"]
                for ci in range(len(raw_cells)):
                    val = raw_cells[ci]
                    if val is not None and isinstance(val, (int, float)) and abs(val) < 1:
                        if ci >= 6 and ci <= 8:
                            fy_idx = ci - 6
                            if fy_idx < len(fy_names):
                                op_or_vals[fy_names[fy_idx]] = _safe_pct(val)
                        if ci == 9:
                            weighted = _safe_pct(val)
                if not weighted:
                    all_floats = [v for v in raw_cells if v is not None and isinstance(v, (int, float)) and abs(v) < 1]
                    weighted = _safe_pct(all_floats[-1]) if all_floats else "N/A"
                if data.quartiles:
                    data.quartiles.upper_quartile = weighted
                    data.quartiles.per_year_upper = op_or_vals

    # Build comparable companies from margin analysis
    if data.margin_analysis:
        for ma in data.margin_analysis:
            data.comparable_companies.append(ComparableCompany(
                sno=ma.sno,
                name=ma.name,
                margins=ma.margin_revenue,
                weighted_avg=ma.weighted_margin
            ))

    if not data.fy_years:
        data.fy_years = ["FY 2022", "FY 2023", "FY 2024"]

    if data.total_accepted == 0 and data.comparable_companies:
        data.total_accepted = len(data.comparable_companies)

    wb.close()
    return data


def parse_annexure3(file_path: str) -> dict:
    """
    Parse Annexure 3 (tested party margins) for financial data.
    Returns a dict with financial values.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    financials = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [_safe_str(c) for c in row]
        if not any(cells):
            continue

        label = cells[0].lower() if cells[0] else ""
        value = row[1] if len(row) > 1 else None

        if "operating revenue" == label and value is not None:
            financials["operating_revenue"] = float(value)
        elif "cost of sales" in label and value is not None:
            financials["cost_of_sales"] = float(value)
        elif "admin" in label and value is not None:
            financials["admin_expenses"] = float(value)
        elif "other expenses" in label and value is not None:
            financials["other_expenses"] = float(value)
        elif "staff salary" in label and value is not None:
            financials["staff_salary"] = float(value)
        elif "partners salaries" in label or "partners sal" in label:
            if value is not None:
                financials["partner_salaries"] = float(value)
        elif "op/or" in label and value is not None:
            financials["op_or"] = float(value)
        elif "op/oc" in label and value is not None:
            financials["op_oc"] = float(value)

    wb.close()
    return financials
